"""
utils/abap_review_exporter.py
Export ABAP code review findings to Excel (colour-coded) and CSV.
"""

import io
import csv
from datetime import datetime
from typing import Any


# ── Excel export (openpyxl) ──────────────────────────────────────────────────

_PRIORITY_FILL = {
    "HIGH":   {"fgColor": "FFD9534F"},   # red
    "MEDIUM": {"fgColor": "FFF0AD4E"},   # amber
    "LOW":    {"fgColor": "FF5CB85C"},   # green
}

_CATEGORY_FILL = {
    "Performance":       {"fgColor": "FFCCE5FF"},
    "Security":          {"fgColor": "FFFFCCCC"},
    "Clean Code":        {"fgColor": "FFCCFFCC"},
    "SOLID":             {"fgColor": "FFFFEECC"},
    "SAP Best Practice": {"fgColor": "FFE8CCFF"},
    "Logic Error":       {"fgColor": "FFFFE0E0"},
    "Naming":            {"fgColor": "FFE0F0FF"},
}


def export_findings_to_excel(
    findings: list[dict],
    class_name: str = "ABAP Class",
    review_mode: str = "Quick",
    focus: str = "General Review",
) -> bytes:
    """Return an Excel workbook as bytes."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Code Review Findings"

    # ── Header section ───────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="FF0A6ED1")
    white_bold = Font(bold=True, color="FFFFFFFF", size=11)
    label_font = Font(bold=True, size=10)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"ABAP Code Review — {class_name}"
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(bold=True, color="FFFFFFFF", size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("B2", "Class / Object:", "C2", class_name),
        ("B3", "Review Mode:",    "C3", review_mode),
        ("B4", "Focus Area:",     "C4", focus),
        ("B5", "Generated:",      "C5", now),
        ("B6", "Total Findings:", "C6", len(findings)),
    ]
    for lc, lv, vc, vv in meta:
        ws[lc] = lv
        ws[lc].font = label_font
        ws[vc] = vv

    # Priority summary
    counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for f in findings:
        counts[f.get("priority", "LOW")] = counts.get(f.get("priority", "LOW"), 0) + 1

    ws["E2"] = "HIGH"
    ws["F2"] = counts["HIGH"]
    ws["E2"].fill = PatternFill("solid", fgColor="FFD9534F")
    ws["E2"].font = Font(bold=True, color="FFFFFFFF")

    ws["E3"] = "MEDIUM"
    ws["F3"] = counts["MEDIUM"]
    ws["E3"].fill = PatternFill("solid", fgColor="FFF0AD4E")
    ws["E3"].font = Font(bold=True, color="FFFFFFFF")

    ws["E4"] = "LOW"
    ws["F4"] = counts["LOW"]
    ws["E4"].fill = PatternFill("solid", fgColor="FF5CB85C")
    ws["E4"].font = Font(bold=True, color="FFFFFFFF")

    ws.row_dimensions[7].height = 6   # spacer

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["#", "Method / Scope", "Category", "Priority", "Finding", "Example / Fix"]
    col_widths = [4, 22, 20, 10, 55, 55]
    header_row = 8

    col_fill = PatternFill("solid", fgColor="FF003F7F")
    col_font = Font(bold=True, color="FFFFFFFF", size=10)

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.fill = col_fill
        cell.font = col_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[header_row].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Sort: HIGH first, then MEDIUM, then LOW
    priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    sorted_findings = sorted(
        findings,
        key=lambda f: (priority_order.get(f.get("priority", "LOW"), 3),
                       f.get("method", ""))
    )

    for ri, finding in enumerate(sorted_findings, start=header_row + 1):
        priority = finding.get("priority", "LOW")
        category = finding.get("category", "")
        p_fill_cfg = _PRIORITY_FILL.get(priority, {"fgColor": "FFFFFFFF"})
        p_fill = PatternFill("solid", **p_fill_cfg)
        c_fill_cfg = _CATEGORY_FILL.get(category, {"fgColor": "FFFFFFFF"})
        c_fill = PatternFill("solid", **c_fill_cfg)

        row_bg = "FFF9F9F9" if ri % 2 == 0 else "FFFFFFFF"
        bg_fill = PatternFill("solid", fgColor=row_bg)

        values = [
            ri - header_row,
            finding.get("method", "GLOBAL"),
            category,
            priority,
            finding.get("suggestion", ""),
            finding.get("example", ""),
        ]
        fills = [bg_fill, bg_fill, c_fill, p_fill, bg_fill, bg_fill]
        fonts = [
            Font(size=9, color="FF888888"),
            Font(bold=True, size=10),
            Font(size=10),
            Font(bold=True, size=10, color="FFFFFFFF"),
            Font(size=10),
            Font(size=9, color="FF444444"),
        ]

        for ci, (val, fill, font) in enumerate(zip(values, fills, fonts), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = wrap

        ws.row_dimensions[ri].height = 55

    # Freeze panes below header
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_findings_to_csv(findings: list[dict]) -> bytes:
    """Return findings as UTF-8 CSV bytes."""
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=["method", "category", "priority", "suggestion", "example"],
        extrasaction="ignore",
    )
    writer.writeheader()
    for f in findings:
        writer.writerow(f)
    return buf.getvalue().encode("utf-8")
